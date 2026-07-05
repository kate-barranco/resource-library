#!/usr/bin/env python3
"""
Extracts the raw Airtable/Excel export into normalized JSON.

Source: Downloads/Social Change Toolbox - Resource Collection - RESOURCE LIBRARY FOR WEBSITE (2).xlsx
Output:
  data/raw/live.json          -- resources with a working URL (goes into the site)
  data/raw/needs_review.json  -- resources with NO discoverable URL (for Kate to review/source manually)
  data/raw/excluded.json      -- rows from directory-style sheets excluded from this build (kept for reference)

Run: python3 scripts/extract.py
"""
import json
import re
import openpyxl
from pathlib import Path

SRC = "/Users/katebarranco/Downloads/Social Change Toolbox - Resource Collection - RESOURCE LIBRARY FOR WEBSITE (2).xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
OUT_DIR.mkdir(parents=True, exist_ok=True)

URL_RE = re.compile(r'https?://\S+')

# Sheets excluded entirely from this build (directories, not discoverable resources) -- confirmed with Kate.
EXCLUDED_SHEETS = {
    "Organizations x Social Change",
    "Funders x Social Change",
    "CertificatesCourses",
    "Job Process Resources",
    "Spec. Contacts x Social Change",
}

# Raw tab name -> clean category. Confirmed with Kate.
TAB_TO_CATEGORY = {
    "Systems ChangeSystems DesignFut": "Systems Change, Design & Futures Thinking",
    "Future-Facing Policy InnovVisua": "Systems Change, Design & Futures Thinking",
    "Polymeta-crisis  Societal Colla": "Systems Change, Design & Futures Thinking",
    "Theories of ChangeDesign Logics": "Theories of Change & Impact Mapping",
    "Impact Mapping": "Theories of Change & Impact Mapping",
    "Social Innovation x Entrepreneu": "Social Innovation & Entrepreneurship",
    "Storytelling x Social Change": "Storytelling for Social Change",
    "Misc. Interesting Articles": "Storytelling for Social Change",
    "Imagination x Dreaming  Play x ": "Imagination, Play & Creativity",
    "LiteracyPost-Literacy Society": "Imagination, Play & Creativity",
    "Hope  Empowerment": "Hope & Empowerment",
    "RelationalityConnectn x Grounde": "Relationality, Connection & Community Health",
    "LonelinessCommunity HealthMens ": "Relationality, Connection & Community Health",
    "Indiv. Healing Practice  Applie": "Individual Healing Practice & Applied Somatics",
    "Org WB x Embodied Leadership": "Organizational Wellbeing & Embodied Leadership",
    "Global Mental Health RESEARCH (": "Global Mental Health Research",
    "Mental Health RESOURCEWEBSITES ": "Global Mental Health Research",
    "Intergenerational TraumaTrauma-": "Intergenerational Trauma & Trauma-Informed Practice",
    "Trauma-Informed in Practice": "Intergenerational Trauma & Trauma-Informed Practice",
    "IGTMH x ArtLiterature x Healing": "Art & Healing",
    "GenZ-Exp, MH, & Identity  Yth C": "Gen Z, Youth Experience & Identity",
    "SM x Youth x MH x Attention Eco": "Social Media, Attention Economy & Youth MH",
    "Tech x Social Change": "Tech, AI & Social Change",
    "AI x WB": "Tech, AI & Social Change",
    "Economy x Philanthropy x WB x E": "Economy, Philanthropy & Wellbeing",
    "Spirit. Ecology x EB x M-Nature": "Spiritual Ecology & Ecological Belonging",
    "Climate ChangeJustice x Eco-Anx": "Climate Justice & Eco-Anxiety",
    "Govt x Demo x War x Polarz  x H": "Government, Democracy & Polarization",
    "Racial Equity & Anti-Racism & D": "Racial Equity, Reproductive Justice & Gender",
    "Reproductive Justice x GenderSe": "Racial Equity, Reproductive Justice & Gender",
    "Gun Violence": "Racial Equity, Reproductive Justice & Gender",
}

# Heuristic type normalization -- raw values are extremely inconsistent (109 distinct strings).
TYPE_RULES = [
    (r'\bpodcast\b', "Podcast"),
    (r'\bted ?talk\b|\btalk\b|\bvideo\b|\bdocumentary\b|\bfilm\b', "Video/Talk"),
    (r'\bbook\b', "Book"),
    (r'\bacademic\b|\bjournal article\b|\bpeer.review', "Academic Paper"),
    (r'\breport\b|\blandscape analysis\b|\bbriefing\b|\bwhite ?paper\b', "Report"),
    (r'\bstudy\b|\bresearch\b|\bsurvey\b|\bpoll\b', "Study/Research"),
    (r'\btoolkit\b|\btool\b|\bframework\b|\bexercise\b|\bcard', "Tool/Framework"),
    (r'\bhub\b|\bsheet\b|\blist\b|\bwebsite\b|\borg(anization)?\b|\bresource doc\b', "Resource Hub"),
    (r'\binterview\b', "Interview"),
    (r'\bblog\b|\barticle\b|\bonline\b|\bnews\b|\breporting\b|\bfeature\b|\bmedium\b', "Article"),
]

def normalize_type(raw):
    if not raw:
        return "Other"
    s = str(raw).strip().lower()
    if not s:
        return "Other"
    for pattern, label in TYPE_RULES:
        if re.search(pattern, s):
            return label
    return "Other"

def find_header_col(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword.lower() in str(h).lower():
            return i
    return None

def cell_text(v):
    if v is None:
        return ""
    return str(v).strip()

def extract_url(row_cells):
    """Prefer hyperlink on the title cell (col 0), then any hyperlink in the row,
    then any URL-shaped text anywhere in the row."""
    if row_cells[0].hyperlink and row_cells[0].hyperlink.target:
        return row_cells[0].hyperlink.target.strip()
    for c in row_cells:
        if c.hyperlink and c.hyperlink.target:
            return c.hyperlink.target.strip()
    for c in row_cells:
        if isinstance(c.value, str):
            m = URL_RE.search(c.value)
            if m:
                return m.group(0).strip()
    return None

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    live, needs_review, excluded = [], [], []
    counter = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        type_col = find_header_col(headers, "type")
        note_col = find_header_col(headers, "note") or find_header_col(headers, "relevance")
        tag_col = find_header_col(headers, "additional tag") or find_header_col(headers, "tag")
        hub_col = find_header_col(headers, "hub")
        year_col = find_header_col(headers, "year")
        source_col = find_header_col(headers, "source")

        excluded_sheet = sn in EXCLUDED_SHEETS
        category = TAB_TO_CATEGORY.get(sn, sn)

        for r in range(2, ws.max_row + 1):
            row_cells = ws[r]
            vals = [c.value for c in row_cells]
            if not any(v not in (None, "") for v in vals):
                continue

            title = cell_text(vals[0])
            if not title:
                continue

            url = extract_url(row_cells)
            if not url and source_col is not None:
                url = None  # already checked via extract_url which scans all cells incl. source_col text
            raw_type = cell_text(vals[type_col]) if type_col is not None else ""
            note = cell_text(vals[note_col]) if note_col is not None else ""
            extra_tag = cell_text(vals[tag_col]) if tag_col is not None else ""
            hub = cell_text(vals[hub_col]) if hub_col is not None else ""
            year = cell_text(vals[year_col]) if year_col is not None else ""

            counter += 1
            record = {
                "raw_id": f"RAW-{counter:04d}",
                "title": title,
                "url": url,
                "sheet": sn,
                "category": category,
                "type_raw": raw_type,
                "type_normalized": normalize_type(raw_type),
                "note": note,
                "extra_tag_text": extra_tag,
                "hub": hub,
                "year": year,
            }

            if excluded_sheet:
                excluded.append(record)
            elif url:
                live.append(record)
            else:
                needs_review.append(record)

    (OUT_DIR / "live.json").write_text(json.dumps(live, indent=2, ensure_ascii=False))
    (OUT_DIR / "needs_review.json").write_text(json.dumps(needs_review, indent=2, ensure_ascii=False))
    (OUT_DIR / "excluded.json").write_text(json.dumps(excluded, indent=2, ensure_ascii=False))

    print(f"live: {len(live)}")
    print(f"needs_review: {len(needs_review)}")
    print(f"excluded: {len(excluded)}")

    cats = {}
    for r in live:
        cats[r["category"]] = cats.get(r["category"], 0) + 1
    print("\nLive resources by category:")
    for c, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {c}")

if __name__ == "__main__":
    main()
